#!/usr/bin/env python3
"""
LongevityPath Evidence Registry — JSON-backed CLI

Manages the evidence map: studies, claims, study↔claim links, and evidence usage.
All data stored in studies.json (no SQLite, works on any filesystem).

DATA SEPARATION: This file stores ONLY scientific evidence metadata.
User health data stays client-side in localStorage. See ARCHITECTURE.md §3.6.

Usage:
    python registry.py init                       Create empty studies.json
    python registry.py import-md FILE             Parse markdown → studies.json
    python registry.py import-refs PAGE           Parse studyRefs from evidence JSON → studies.json
    python registry.py query CLAIM [--dir +/-/±]  Studies for a claim
    python registry.py summary                    Claim Summary Index
    python registry.py export-summary             Write index to study-registry.md
    python registry.py export-refs CLAIM [--json] Builder-ready study refs
    python registry.py export-xlsx [CATEGORY]     Export studies to xlsx for review
    python registry.py stale                      Studies verified >6 months ago
    python registry.py gaps [--category CAT]      Claims with weak or missing evidence
    python registry.py supersede OLD_ID NEW_ID    Mark a study as superseded
    python registry.py enrich-all                 List studies with incomplete fields
    python registry.py add                        Interactive: add a study
    python registry.py stats                      Database statistics
    python registry.py verify-dois [--category]   Batch-verify all DOIs against doi.org
"""

import argparse
import json
import re
import sys
import uuid
import urllib.request
import urllib.error
from datetime import datetime, timedelta
from pathlib import Path

DATA_PATH = Path(__file__).parent / "studies.json"
REGISTRY_MD = Path(__file__).parent / "study-registry.md"


# ─────────────────────────────────────────────
# Data access
# ─────────────────────────────────────────────

def load_db():
    """Load studies.json, creating empty structure if missing."""
    if DATA_PATH.exists():
        return json.loads(DATA_PATH.read_text(encoding='utf-8'))
    return {'studies': [], 'claims': [], 'study_claims': [], 'evidence_usage': []}


def save_db(db):
    """Write studies.json."""
    DATA_PATH.write_text(json.dumps(db, indent=2, ensure_ascii=False), encoding='utf-8')


def uid():
    return uuid.uuid4().hex[:12]


# ─────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────

def find_studies_for_claim(db, claim_id, direction=None):
    """Return studies linked to a claim, sorted by final_score desc.
    Each returned study dict gets an extra 'direction' key from the link."""
    # Build map: study_id → direction from the link
    link_map = {lk['study_id']: lk.get('direction', '')
                for lk in db['study_claims'] if lk['claim_id'] == claim_id}
    results = []
    for s in db['studies']:
        if s['study_id'] in link_map:
            enriched = dict(s)
            enriched['direction'] = link_map[s['study_id']]
            results.append(enriched)
    if direction:
        results = [s for s in results if s.get('direction') == direction]
    results.sort(key=lambda s: s.get('final_score', 0), reverse=True)
    return results


def make_study_id(authors, year):
    author = authors.strip().strip('*').split(' ')[0].split('&')[0].strip()
    author = re.sub(r'[^a-zA-Z]', '', author).lower()
    return f"{author}-{year}-{uid()[:4]}"


# ─────────────────────────────────────────────
# DOI validation
# ─────────────────────────────────────────────

def validate_doi(doi, expected_title=None, expected_authors=None, timeout=10):
    """Validate a DOI by resolving it via doi.org content negotiation.

    Returns dict with:
        valid: bool - DOI resolves to a real paper
        title_match: bool|None - title matches if expected_title given
        resolved_title: str - title from DOI metadata
        resolved_authors: str - first author from DOI metadata
        error: str|None - error message if validation failed
    """
    result = {'valid': False, 'title_match': None, 'resolved_title': '',
              'resolved_authors': '', 'error': None}

    if not doi:
        result['error'] = 'No DOI provided'
        return result

    # Clean DOI
    doi = doi.strip()
    if doi.startswith('https://doi.org/'):
        doi = doi[len('https://doi.org/'):]
    if doi.startswith('http://doi.org/'):
        doi = doi[len('http://doi.org/'):]

    url = f'https://doi.org/{doi}'
    req = urllib.request.Request(url, headers={
        'Accept': 'application/vnd.citationstyles.csl+json',
        'User-Agent': 'LongevityPath-Registry/1.0 (mailto:registry@longevitypath.org)'
    })

    try:
        with urllib.request.urlopen(req, timeout=timeout) as resp:
            data = json.loads(resp.read().decode('utf-8'))
    except urllib.error.HTTPError as e:
        if e.code == 404:
            result['error'] = f'DOI not found (404): {doi}'
        else:
            result['error'] = f'HTTP {e.code} resolving DOI: {doi}'
        return result
    except urllib.error.URLError as e:
        result['error'] = f'Network error resolving DOI: {e.reason}'
        return result
    except Exception as e:
        result['error'] = f'Error resolving DOI: {str(e)}'
        return result

    result['valid'] = True

    # Extract title
    resolved_title = data.get('title', '')
    if isinstance(resolved_title, list):
        resolved_title = resolved_title[0] if resolved_title else ''
    result['resolved_title'] = resolved_title

    # Extract first author
    authors = data.get('author', [])
    if authors:
        first = authors[0]
        result['resolved_authors'] = f"{first.get('family', '')} {first.get('given', '')}".strip()

    # Title comparison (fuzzy — normalize and compare first 40 chars)
    if expected_title:
        def normalize(t):
            return re.sub(r'[^a-z0-9]', '', t.lower())[:60]
        n_expected = normalize(expected_title)
        n_resolved = normalize(resolved_title)
        # Check if they share significant overlap
        result['title_match'] = (n_expected[:40] == n_resolved[:40]) or (n_expected in n_resolved) or (n_resolved in n_expected)

    return result


def validate_study_doi(study, strict=True):
    """Validate a study's DOI and check title match.

    Returns (is_ok: bool, message: str)
    """
    doi = study.get('doi', '')
    if not doi:
        if strict:
            return False, f"  ✗ {study['study_id']}: No DOI"
        return True, f"  ⚠ {study['study_id']}: No DOI (skipped)"

    title = study.get('title', '')
    result = validate_doi(doi, expected_title=title)

    if not result['valid']:
        return False, f"  ✗ {study['study_id']}: {result['error']}"

    if result['title_match'] is False:
        return False, (
            f"  ✗ {study['study_id']}: TITLE MISMATCH\n"
            f"    Expected: {title[:80]}\n"
            f"    Resolved: {result['resolved_title'][:80]}\n"
            f"    DOI: {doi}"
        )

    return True, f"  ✓ {study['study_id']}: DOI verified — {result['resolved_title'][:60]}"


# ─────────────────────────────────────────────
# verify-dois
# ─────────────────────────────────────────────

def cmd_verify_dois(args):
    """Batch-verify all DOIs in studies.json against doi.org."""
    import time

    db = load_db()
    category = getattr(args, 'category', None)

    # Filter by category if requested
    if category:
        claim_ids = {c['claim_id'] for c in db['claims']
                     if get_claim_category(c['claim_id']) == category}
        study_ids = {sc['study_id'] for sc in db['study_claims']
                     if sc['claim_id'] in claim_ids}
        studies = [s for s in db['studies'] if s['study_id'] in study_ids]
    else:
        studies = db['studies']

    total = len(studies)
    ok = 0
    failed = []
    no_doi = []
    mismatched = []

    print(f"\n  Verifying {total} study DOIs against doi.org...\n")

    for i, s in enumerate(studies, 1):
        sid = s['study_id']
        doi = s.get('doi', '')

        if not doi:
            no_doi.append(s)
            print(f"  [{i}/{total}] ⚠ {sid}: No DOI")
            continue

        # Rate-limit: ~1 request per second to be polite
        if i > 1:
            time.sleep(1.0)

        is_ok, msg = validate_study_doi(s, strict=True)
        print(f"  [{i}/{total}] {msg.strip()}")

        if is_ok:
            ok += 1
        else:
            if 'TITLE MISMATCH' in msg:
                mismatched.append((s, msg))
            else:
                failed.append((s, msg))

    # Summary
    print(f"\n{'='*60}")
    print(f"  DOI Verification Summary")
    print(f"{'='*60}")
    print(f"  Total studies:   {total}")
    print(f"  DOI verified OK: {ok}")
    print(f"  No DOI:          {len(no_doi)}")
    print(f"  Failed/404:      {len(failed)}")
    print(f"  Title mismatch:  {len(mismatched)}")

    if failed:
        print(f"\n  ✗ FAILED DOIs:")
        for s, msg in failed:
            print(f"    {s['study_id']}: {s.get('doi', '')} — {s.get('authors', '')} {s.get('pub_year', '')}")

    if mismatched:
        print(f"\n  ⚠ TITLE MISMATCHES (DOI points to wrong paper?):")
        for s, msg in mismatched:
            print(f"    {s['study_id']}: {s.get('doi', '')}")
            print(f"      Expected: {s.get('title', '')[:80]}")

    if not failed and not mismatched:
        print(f"\n  ✓ All DOIs verified successfully!")

    print()


# ─────────────────────────────────────────────
# init
# ─────────────────────────────────────────────

def cmd_init(args):
    if DATA_PATH.exists():
        print(f"studies.json already exists ({DATA_PATH})")
        return
    save_db({'studies': [], 'claims': [], 'study_claims': [], 'evidence_usage': []})
    print(f"✓ Created {DATA_PATH}")


# ─────────────────────────────────────────────
# import-md
# ─────────────────────────────────────────────

SECTION_PATTERN = re.compile(r'^## (.+)$')
TABLE_ROW_PATTERN = re.compile(r'^\|(.+)\|$')

SECTION_TO_PAGE = {
    "Sleep": "sleep", "Mindset": "mindset", "Wellbeing": "wellbeing",
    "Nutrition": "nutrition", "Protein": "protein", "VO2max": "vo2max",
    "Muscle Strength": "muscle",
}


def parse_claim_tags(claims_str):
    return re.findall(r'`([^`]+→[^`]+)`', claims_str)


def parse_used_in(used_in_str, section_page):
    usages = []
    if not used_in_str or used_in_str.strip() == '—':
        return usages
    parts = re.split(r',\s*(?=[a-z])', used_in_str.strip())
    for part in parts:
        part = part.strip()
        if not part:
            continue
        role_match = re.search(r'\(([FS])\)', part)
        role = 'featured' if role_match and role_match.group(1) == 'F' else 'supporting'
        part_clean = re.sub(r'\([FS]\)', '', part).strip()
        page_match = re.match(r'([a-z0-9]+)#(.+)', part_clean)
        if page_match:
            page = page_match.group(1)
            cards = re.findall(r'q\d+', page_match.group(2))
            page_file = f"{page}-evidence.html"
            for card in cards:
                usages.append((page_file, card, role))
        else:
            cards = re.findall(r'q\d+', part_clean)
            page_file = f"{section_page}-evidence.html" if section_page else "unknown.html"
            for card in cards:
                usages.append((page_file, card, role))
    return usages


def parse_score(score_str):
    try:
        final = float(score_str.strip())
    except ValueError:
        return (0.0, 1.0, 0.0)
    if final == int(final) and final <= 14:
        return (final, 1.0, final)
    elif final < 7:
        return (final * 2, 0.5, final)
    return (final, 1.0, final)


def cmd_import_md(args):
    md_path = Path(args.file)
    if not md_path.exists():
        print(f"✗ File not found: {md_path}")
        sys.exit(1)

    db = load_db()
    existing_dois = {s['doi'] for s in db['studies'] if s.get('doi')}
    existing_claims = {c['claim_id'] for c in db['claims']}
    text = md_path.read_text(encoding='utf-8')
    lines = text.splitlines()

    studies_added = 0

    # Extract claim vocabulary
    in_vocab = False
    for line in lines:
        if '## Claim Tag Vocabulary' in line:
            in_vocab = True
            continue
        if in_vocab and line.startswith('## '):
            break
        if in_vocab and line.startswith('|'):
            match = TABLE_ROW_PATTERN.match(line)
            if match:
                cells = [c.strip() for c in match.group(1).split('|')]
                if len(cells) >= 2:
                    tag = cells[0].strip().strip('`')
                    desc = cells[1].strip()
                    if '→' in tag and tag != 'Tag' and tag not in existing_claims:
                        parts = tag.split('→', 1)
                        db['claims'].append({
                            'claim_id': tag, 'exposure': parts[0],
                            'outcome': parts[1], 'description': desc
                        })
                        existing_claims.add(tag)

    # Parse study tables
    current_section = None
    current_page = None
    in_removed = False

    for line in lines:
        sec_match = SECTION_PATTERN.match(line)
        if sec_match:
            name = sec_match.group(1).strip()
            if name in SECTION_TO_PAGE:
                current_section = name
                current_page = SECTION_TO_PAGE[name]
                in_removed = False
            elif 'Removed' in name:
                in_removed = True
                current_section = None
            else:
                current_section = None
                in_removed = False
            continue

        if not line.startswith('|') or in_removed or not current_section:
            continue

        match = TABLE_ROW_PATTERN.match(line)
        if not match:
            continue
        cells = [c.strip() for c in match.group(1).split('|')]
        if not cells or cells[0].startswith('---') or cells[0] == 'Study':
            continue
        if len(cells) < 12:
            continue

        authors = cells[0].strip('*').strip()
        try:
            pub_year = 2026 - int(cells[1].strip())
        except ValueError:
            pub_year = 2020
        doi = cells[2].strip() if cells[2].strip().startswith('10.') else None
        study_type = cells[3].strip()
        sample = cells[4].strip() or None
        q, r, f = parse_score(cells[5])
        landmark = cells[6].strip().upper() == 'Y'
        direction = {'+': '+', '−': '−', '-': '−', '±': '±'}.get(cells[7].strip(), cells[7].strip())
        population = cells[8].strip() or 'all'
        claims_str = cells[9].strip()
        used_in_str = cells[10].strip()
        notes = cells[11].strip() if len(cells) > 11 else ''

        if doi and doi in existing_dois:
            continue

        sid = make_study_id(authors, pub_year)
        db['studies'].append({
            'study_id': sid, 'authors': authors, 'pub_year': pub_year,
            'doi': doi, 'study_type': study_type, 'sample_size': sample,
            'quality_score': q, 'relevance_mult': r, 'final_score': f,
            'landmark': landmark, 'direction': direction, 'population': population,
            'key_finding': notes or None, 'verified_date': '2026-02',
            'notes': notes or None,
        })
        if doi:
            existing_dois.add(doi)
        studies_added += 1

        for tag in parse_claim_tags(claims_str):
            if tag not in existing_claims:
                parts = tag.split('→', 1)
                db['claims'].append({'claim_id': tag, 'exposure': parts[0], 'outcome': parts[1], 'description': ''})
                existing_claims.add(tag)
            link = {'study_id': sid, 'claim_id': tag}
            if link not in db['study_claims']:
                db['study_claims'].append(link)

        for page_file, card_id, role in parse_used_in(used_in_str, current_page):
            db['evidence_usage'].append({
                'id': uid(), 'study_id': sid, 'page_file': page_file,
                'card_id': card_id, 'role': role
            })

    save_db(db)
    print(f"✓ Import complete: {studies_added} studies, {len(existing_claims)} claims")


# ─────────────────────────────────────────────
# query
# ─────────────────────────────────────────────

def cmd_query(args):
    db = load_db()
    results = find_studies_for_claim(db, args.claim, args.dir)

    if not results:
        fuzzy = [c['claim_id'] for c in db['claims'] if args.claim in c['claim_id']]
        if fuzzy:
            print(f"No exact match for '{args.claim}'. Did you mean:")
            for c in fuzzy:
                print(f"  • {c}")
        else:
            print(f"No studies found for claim '{args.claim}'")
        return

    print(f"\n{'='*70}")
    print(f"  Claim: {args.claim}  ({len(results)} studies)")
    print(f"{'='*70}\n")

    for s in results:
        lm = " ★" if s.get('is_landmark') else ""
        pop = f" [{s['population']}]" if s.get('population', 'all') != 'all' else ""
        print(f"  {s.get('direction','')}  {s['authors']} {s['pub_year']}  "
              f"({s.get('study_type','')}, Score {s.get('final_score',0):.0f}{lm}){pop}")
        if s.get('key_finding'):
            print(f"     → {s['key_finding'][:100]}")
        if s.get('doi'):
            print(f"     DOI: {s['doi']}")
        print()


# ─────────────────────────────────────────────
# summary
# ─────────────────────────────────────────────

def build_summary_data(db):
    summary = []
    claim_ids = sorted({c['claim_id'] for c in db['claims']})
    for cid in claim_ids:
        studies = find_studies_for_claim(db, cid)
        if not studies:
            continue

        plus = [s for s in studies if s.get('direction') == '+']
        minus = [s for s in studies if s.get('direction') == '−']
        mixed = [s for s in studies if s.get('direction') == '±']

        best_plus = f"{plus[0]['authors']} {plus[0]['pub_year']} ({plus[0]['final_score']:.0f})" if plus else '—'
        best_minus = f"{minus[0]['authors']} {minus[0]['pub_year']} ({minus[0]['final_score']:.0f})" if minus else '—'

        if plus and not minus and not mixed:
            net = '+'
        elif minus and not plus:
            net = '−'
        elif mixed and not plus and not minus:
            net = '±'
        else:
            bp = plus[0]['final_score'] if plus else 0
            bm = minus[0]['final_score'] if minus else 0
            net = '+ (contested)' if bp >= bm else '− (contested)'

        top = max(s.get('final_score', 0) for s in studies)
        confidence = 'Strong' if top >= 12 else 'Moderate' if top >= 10 else 'Limited'

        gap = ''
        if not minus and not mixed and plus:
            gap = 'Need contradicting study'
        elif minus and not plus:
            gap = 'Need supporting study'

        summary.append({
            'claim': cid, 'n_plus': len(plus), 'n_minus': len(minus),
            'n_mixed': len(mixed), 'best_plus': best_plus, 'best_minus': best_minus,
            'net': net, 'confidence': confidence, 'gap': gap,
        })
    return summary


def cmd_summary(args):
    db = load_db()
    summary = build_summary_data(db)
    if not summary:
        print("No claims with studies. Run 'import-md' first.")
        return

    print(f"\n{'='*90}")
    print("  Claim Summary Index")
    print(f"{'='*90}\n")
    print(f"  {'Claim':<35} {'#+':>3} {'#−':>3} {'#±':>3}  {'Net':<16} {'Confidence':<12} {'Gap'}")
    print(f"  {'─'*35} {'─'*3} {'─'*3} {'─'*3}  {'─'*16} {'─'*12} {'─'*25}")
    for s in summary:
        print(f"  {s['claim']:<35} {s['n_plus']:>3} {s['n_minus']:>3} {s['n_mixed']:>3}  "
              f"{s['net']:<16} {s['confidence']:<12} {s['gap']}")


# ─────────────────────────────────────────────
# export-summary
# ─────────────────────────────────────────────

def cmd_export_summary(args):
    db = load_db()
    summary = build_summary_data(db)
    if not summary:
        print("No claims with studies.")
        return

    lines = [
        "| Claim | #+ | #− | #± | Best+ | Best− | Net | Confidence | Gap? |",
        "|-------|----|----|-----|-------|-------|-----|------------|------|"
    ]
    for s in summary:
        lines.append(
            f"| `{s['claim']}` | {s['n_plus']} | {s['n_minus']} | {s['n_mixed']} | "
            f"{s['best_plus']} | {s['best_minus']} | {s['net']} | {s['confidence']} | {s['gap']} |"
        )

    text = REGISTRY_MD.read_text(encoding='utf-8')
    pattern = re.compile(r'(\| Claim \| #\+ \|.*?\n\|[-|\s]+\n)((?:\|.*\n)*)', re.MULTILINE)
    match = pattern.search(text)
    if match:
        text = text[:match.start()] + '\n'.join(lines) + '\n' + text[match.end():]
        REGISTRY_MD.write_text(text, encoding='utf-8')
        print(f"✓ Updated study-registry.md ({len(summary)} claims)")
    else:
        print("✗ Could not find Claim Summary Index table")


# ─────────────────────────────────────────────
# export-refs
# ─────────────────────────────────────────────

def cmd_export_refs(args):
    db = load_db()

    if args.claim == 'all':
        # Group by claim
        claim_ids = sorted({lk['claim_id'] for lk in db['study_claims']})
        rows = []
        for cid in claim_ids:
            for s in find_studies_for_claim(db, cid):
                rows.append({**s, 'claim_id': cid})
    else:
        studies = find_studies_for_claim(db, args.claim)
        rows = [{**s, 'claim_id': args.claim} for s in studies]

    if not rows:
        print(f"No studies found for claim: {args.claim}", file=sys.stderr)
        sys.exit(1)

    if args.json:
        out = []
        for r in rows:
            out.append({
                'authors': r['authors'], 'year': r['pub_year'], 'doi': r.get('doi'),
                'studyType': r.get('study_type'), 'sampleSize': r.get('sample_size'),
                'qualityScore': r.get('quality_score'), 'keyFinding': r.get('key_finding'),
                'claim': r['claim_id'],
            })
        print(json.dumps(out, indent=2))
    else:
        current_claim = None
        ref_num = 0
        for r in rows:
            if r['claim_id'] != current_claim:
                current_claim = r['claim_id']
                ref_num = 0
                print(f"\n--- {current_claim} ---\n")

            ref_num += 1
            doi_str = f'doi:{r["doi"]}' if r.get('doi') else 'no DOI'
            badge = r.get('study_type', 'Study')

            if ref_num == 1:
                print(f'<!-- FEATURED (score={r.get("final_score", 0)}) -->')
                print(f'<div class="study-citation">')
                print(f'  <div class="study-header">')
                print(f'    <span class="study-badge meta-analysis">{badge}</span>')
                print(f'  </div>')
                if r.get('key_finding'):
                    print(f'  <div class="study-finding"><strong>Key finding:</strong> {r["key_finding"]}</div>')
                print(f'  <div class="study-detail">{r["authors"]} {r["pub_year"]} · N={r.get("sample_size") or "?"}</div>')
                print(f'</div>')
                print()

            print(f'<div class="study-ref">[{ref_num}] {r["authors"]} ({r["pub_year"]}). {r.get("key_finding") or ""} {doi_str}</div>')


# ─────────────────────────────────────────────
# stale
# ─────────────────────────────────────────────

def cmd_stale(args):
    db = load_db()
    cutoff = (datetime.now() - timedelta(days=180)).strftime('%Y-%m')
    stale = [s for s in db['studies']
             if not s.get('verified_date') or s['verified_date'] < cutoff]
    stale.sort(key=lambda s: s.get('verified_date') or '')

    if not stale:
        print("✓ No stale studies. All verified within 6 months.")
        return

    print(f"\n⚠ {len(stale)} stale studies (verified before {cutoff}):\n")
    for s in stale:
        print(f"  {s['authors']} {s['pub_year']}  (Score {s.get('final_score',0):.0f})  "
              f"Verified: {s.get('verified_date') or 'never'}")


# ─────────────────────────────────────────────
# add
# ─────────────────────────────────────────────

def cmd_add(args):
    print("\n─── Add New Study ───\n")
    db = load_db()

    authors = input("Authors (e.g., 'Smith et al.'): ").strip()
    pub_year = int(input("Publication year: ").strip())
    doi = input("DOI (or blank): ").strip() or None
    title = input("Title: ").strip() or None

    # ── DOI validation gate ──
    if doi:
        print(f"\n  Validating DOI: {doi} ...")
        result = validate_doi(doi, expected_title=title)
        if not result['valid']:
            print(f"  ✗ DOI INVALID: {result['error']}")
            print(f"  Study NOT added. Fix the DOI and try again.")
            sys.exit(1)
        if result['title_match'] is False:
            print(f"  ⚠ TITLE MISMATCH:")
            print(f"    You entered:  {title[:80]}")
            print(f"    DOI resolves: {result['resolved_title'][:80]}")
            confirm = input("  Continue anyway? (y/N): ").strip().lower()
            if confirm != 'y':
                print("  Study NOT added.")
                sys.exit(0)
        else:
            print(f"  ✓ DOI verified: {result['resolved_title'][:70]}")
        # Auto-fill title from DOI if not provided
        if not title and result['resolved_title']:
            title = result['resolved_title']
            print(f"  Auto-filled title: {title[:70]}")

    study_type = input("Type (MA/SR/RCT/Cohort/etc.): ").strip()
    sample = input("Sample size: ").strip() or None
    quality = float(input("Quality score (0-14): ").strip())
    relevance = float(input("Relevance multiplier (1.0/0.5/0.0): ").strip())
    final = quality * relevance
    landmark = input("Landmark? (Y/N): ").strip().upper() == 'Y'
    direction = input("Direction (+/−/±): ").strip()
    population = input("Population (all/<65/>65/etc.): ").strip() or 'all'
    finding = input("Key finding: ").strip() or None

    print("\nClaim tags (comma-separated, e.g., 'protein→cancer, protein→mortality'):")
    claims_raw = input("> ").strip()
    claim_tags = [c.strip() for c in claims_raw.split(',') if '→' in c]

    sid = make_study_id(authors, pub_year)
    db['studies'].append({
        'study_id': sid, 'authors': authors, 'pub_year': pub_year,
        'doi': doi, 'title': title, 'study_type': study_type, 'sample_size': sample,
        'quality_score': quality, 'relevance_mult': relevance, 'final_score': final,
        'landmark': landmark, 'direction': direction, 'population': population,
        'key_finding': finding, 'verified_date': datetime.now().strftime('%Y-%m'),
        'notes': None,
    })

    existing_claims = {c['claim_id'] for c in db['claims']}
    for tag in claim_tags:
        if tag not in existing_claims:
            parts = tag.split('→', 1)
            db['claims'].append({'claim_id': tag, 'exposure': parts[0], 'outcome': parts[1], 'description': ''})
            existing_claims.add(tag)
        db['study_claims'].append({'study_id': sid, 'claim_id': tag})

    save_db(db)
    print(f"\n✓ Added: {authors} {pub_year} → {sid}")
    print(f"  Claims: {', '.join(claim_tags)}")
    print(f"  Score: {final:.1f} (Q{quality:.0f} × R{relevance}), Dir {direction}")


# ─────────────────────────────────────────────
# stats
# ─────────────────────────────────────────────

def cmd_stats(args):
    db = load_db()
    studies = db['studies']
    # Count directions from study_claims links (not study-level)
    dir_plus = sum(1 for sc in db['study_claims'] if sc.get('direction') == '+')
    dir_minus = sum(1 for sc in db['study_claims'] if sc.get('direction') == '−')
    dir_mixed = sum(1 for sc in db['study_claims'] if sc.get('direction') == '±')
    landmarks = sum(1 for s in studies if s.get('is_landmark'))
    with_doi = sum(1 for s in studies if s.get('doi'))
    with_finding = sum(1 for s in studies if s.get('key_finding'))

    print(f"\n{'='*50}")
    print("  LongevityPath Evidence Registry")
    print(f"{'='*50}\n")
    print(f"  Studies:        {len(studies)}")
    print(f"    With DOI:     {with_doi}")
    print(f"    With finding: {with_finding}")
    print(f"    Supporting:   {dir_plus}")
    print(f"    Contradicting:{dir_minus}")
    print(f"    Conditional:  {dir_mixed}")
    print(f"    Landmarks:    {landmarks}")
    print(f"  Claims:         {len(db['claims'])}")
    print(f"  Study↔Claim:    {len(db['study_claims'])}")
    print(f"  Evidence usage: {len(db['evidence_usage'])}")
    print(f"  Storage:        {DATA_PATH}")
    print()


# ─────────────────────────────────────────────
# gaps
# ─────────────────────────────────────────────

CLAIM_TO_CATEGORY = {}
_CATEGORY_MAP = {
    'sleep': ['sleep-duration', 'sleep-regularity', 'sleep-quality', 'insomnia',
              'long-sleep', 'social-jet-lag', 'catch-up-sleep', 'wearable',
              'caffeine', 'temperature', 'light-therapy', 'warm-bath', 'CBT-I',
              'magnesium', 'ashwagandha', 'glycine', 'melatonin', 'sleep-debt'],
    'nutrition': ['diet-quality', 'fruit-veg', 'UPF', 'mediterranean', 'saturated-fat',
                  'sodium', 'alcohol', 'centenarian-diet', 'flexible-dieting',
                  'eating-speed', 'food-tracking', 'meal-timing'],
    'protein': ['protein', 'protein-source', 'protein-dose', 'protein-distribution',
                'protein-deficit', 'protein-timing', 'amino-acids'],
    'vo2max': ['CRF', 'HIIT', 'exercise-dose', 'PA', 'VO2max-estimation'],
    'muscle-strength': ['muscle-strength', 'resistance-training', 'push-ups', 'chair-stand'],
    'mindset': ['motivation', 'grit', 'locus-of-control', 'growth-mindset', 'mHealth', 'habit'],
    'wellbeing': ['positive-affect', 'flow', 'social-connection', 'purpose', 'goal-pursuit',
                  'stress-mindset', 'emotion-regulation', 'meditation', 'nature', 'gratitude', 'PERMA'],
}
for cat, prefixes in _CATEGORY_MAP.items():
    for prefix in prefixes:
        CLAIM_TO_CATEGORY[prefix] = cat


def get_claim_category(claim_id):
    exposure = claim_id.split('→')[0] if '→' in claim_id else claim_id
    return CLAIM_TO_CATEGORY.get(exposure, 'other')


def cmd_gaps(args):
    db = load_db()
    category = getattr(args, 'category', None)
    summary = build_summary_data(db)

    if category:
        summary = [s for s in summary if get_claim_category(s['claim']) == category]

    weak = [s for s in summary if s['n_plus'] < 2 or s['n_minus'] == 0]
    if not weak:
        print("✓ No evidence gaps found.")
        return

    print(f"\n{'='*80}")
    title = f"  Evidence Gaps" + (f" — {category}" if category else "")
    print(title)
    print(f"{'='*80}\n")
    print(f"  {'Claim':<35} {'#+':>3} {'#−':>3}  {'Confidence':<12} {'Gap'}")
    print(f"  {'─'*35} {'─'*3} {'─'*3}  {'─'*12} {'─'*35}")
    for s in weak:
        gaps = []
        if s['n_plus'] < 2:
            gaps.append(f"Need {2-s['n_plus']} more supporting")
        if s['n_minus'] == 0 and s['n_plus'] > 0:
            gaps.append("Need contradicting study")
        gap_str = '; '.join(gaps)
        print(f"  {s['claim']:<35} {s['n_plus']:>3} {s['n_minus']:>3}  {s['confidence']:<12} {gap_str}")
    print(f"\n  Total gaps: {len(weak)} / {len(summary)} claims\n")


# ─────────────────────────────────────────────
# supersede
# ─────────────────────────────────────────────

def cmd_supersede(args):
    db = load_db()
    old_id, new_id = args.old_id, args.new_id
    old_study = next((s for s in db['studies'] if s['study_id'] == old_id), None)
    new_study = next((s for s in db['studies'] if s['study_id'] == new_id), None)
    if not old_study:
        print(f"✗ Old study not found: {old_id}")
        sys.exit(1)
    if not new_study:
        print(f"✗ New study not found: {new_id}")
        sys.exit(1)
    old_study['status'] = 'superseded'
    old_study['superseded_by'] = new_id
    save_db(db)
    print(f"✓ {old_id} marked as superseded by {new_id}")


# ─────────────────────────────────────────────
# enrich-all
# ─────────────────────────────────────────────

REQUIRED_FIELDS = ['title', 'journal', 'doi', 'study_type', 'key_finding', 'effect_sizes']

def cmd_enrich_all(args):
    db = load_db()
    incomplete = []
    for s in db['studies']:
        missing = []
        for f in REQUIRED_FIELDS:
            val = s.get(f)
            if val is None or val == '' or (isinstance(val, list) and len(val) == 0 and f == 'effect_sizes'):
                missing.append(f)
        if missing:
            incomplete.append((s, missing))

    if not incomplete:
        print("✓ All studies have complete required fields.")
        return

    incomplete.sort(key=lambda x: len(x[1]), reverse=True)
    print(f"\n⚠ {len(incomplete)} studies with incomplete fields:\n")
    for s, missing in incomplete:
        print(f"  {s['authors']:<25} {s['pub_year']}  Missing: {', '.join(missing)}")
    print(f"\n  Required fields: {', '.join(REQUIRED_FIELDS)}")


# ─────────────────────────────────────────────
# export-xlsx
# ─────────────────────────────────────────────

def cmd_export_xlsx(args):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("✗ openpyxl not installed. Run: pip install openpyxl")
        sys.exit(1)

    db = load_db()
    category = getattr(args, 'category', None)

    # Filter studies by category via evidence_usage or study_claims
    if category:
        usage_ids = {eu['study_id'] for eu in db['evidence_usage']
                     if eu.get('evidence_page') == category}
        claim_ids = {c['claim_id'] for c in db['claims']
                     if get_claim_category(c['claim_id']) == category}
        claim_study_ids = {sc['study_id'] for sc in db['study_claims']
                          if sc['claim_id'] in claim_ids}
        target_ids = usage_ids | claim_study_ids
        studies = [s for s in db['studies'] if s['study_id'] in target_ids]
    else:
        studies = db['studies']

    studies.sort(key=lambda s: (-s.get('final_score', 0), s.get('pub_year', 0)))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Evidence {'— ' + category if category else 'All'}"

    headers = [
        'Study ID', 'Authors', 'Year', 'Title', 'Journal', 'DOI',
        'Study Type', 'Sample Size', 'Population', 'Country',
        'Quality Score', 'Risk of Bias', 'GRADE', 'Direction',
        'Status', 'Key Finding', 'Effect Sizes',
        'Claims', 'Evidence Cards', 'Verified', 'Next Review', 'Notes'
    ]

    hfill = PatternFill('solid', fgColor='1B4332')
    hfont = Font(bold=True, color='FFFFFF', size=10, name='Arial')
    dfont = Font(size=9.5, name='Arial')
    wrap = Alignment(wrap_text=True, vertical='top')
    border = Border(
        left=Side(style='thin', color='D0D0D0'), right=Side(style='thin', color='D0D0D0'),
        top=Side(style='thin', color='D0D0D0'), bottom=Side(style='thin', color='D0D0D0')
    )

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = hfont; cell.fill = hfill; cell.alignment = wrap; cell.border = border

    # Build lookups
    study_claims = {}
    for sc in db['study_claims']:
        study_claims.setdefault(sc['study_id'], []).append(
            f"{sc.get('direction', '')} {sc['claim_id']}")
    study_usage = {}
    for eu in db['evidence_usage']:
        study_usage.setdefault(eu['study_id'], []).append(
            f"{eu.get('evidence_page', '')}#{eu.get('card_id', '')}")

    dir_colors = {'+': 'C6F6D5', '−': 'FED7D7', '±': 'FFF3CD'}
    status_colors = {'active': 'FFFFFF', 'superseded': 'F8D7DA', 'retracted': 'F5C6CB'}

    for i, s in enumerate(studies, 1):
        row = i + 1
        effect_str = '; '.join(
            f"{e.get('metric','')} {e.get('value','')} [{e.get('ci_lower','')},{e.get('ci_upper','')}] {e.get('comparison','')[:50]}"
            for e in (s.get('effect_sizes') or [])
        )
        values = [
            s['study_id'], s.get('authors_short') or s.get('authors', ''),
            s.get('pub_year', ''), s.get('title', ''), s.get('journal', ''),
            s.get('doi', ''), s.get('study_type', ''), s.get('sample_size', ''),
            s.get('population', ''), s.get('country', ''),
            s.get('quality_score', ''), s.get('risk_of_bias', ''),
            s.get('grade_certainty', ''), s.get('direction', ''),
            s.get('status', 'active'), s.get('key_finding', ''),
            effect_str,
            ', '.join(study_claims.get(s['study_id'], [])),
            ', '.join(study_usage.get(s['study_id'], [])),
            s.get('verified_date', ''), s.get('next_review_date', ''),
            s.get('notes', ''),
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.font = dfont; cell.alignment = wrap; cell.border = border

        # Color direction
        d = s.get('direction', '')
        if d in dir_colors:
            ws.cell(row=row, column=14).fill = PatternFill('solid', fgColor=dir_colors[d])

        # Color status
        st = s.get('status', 'active')
        if st in status_colors and st != 'active':
            for col in range(1, len(headers) + 1):
                ws.cell(row=row, column=col).fill = PatternFill('solid', fgColor=status_colors[st])

        # Hyperlink DOI
        if s.get('doi'):
            cell = ws.cell(row=row, column=6)
            cell.hyperlink = f"https://doi.org/{s['doi']}"
            cell.font = Font(size=9.5, name='Arial', color='0563C1', underline='single')

    widths = {1:18, 2:18, 3:6, 4:45, 5:22, 6:25, 7:18, 8:12, 9:18, 10:10,
              11:8, 12:12, 13:10, 14:6, 15:10, 16:50, 17:50, 18:30, 19:25, 20:10, 21:10, 22:30}
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f'A1:{get_column_letter(len(headers))}{len(studies)+1}'

    suffix = f'-{category}' if category else ''
    out_path = Path(__file__).parent.parent / f'evidence-registry{suffix}.xlsx'
    wb.save(str(out_path))
    print(f"✓ Exported {len(studies)} studies to {out_path}")


# ─────────────────────────────────────────────
# import-refs
# ─────────────────────────────────────────────

def cmd_import_refs(args):
    evidence_dir = Path(__file__).parent / 'evidence-pages'
    page_name = args.page
    json_path = evidence_dir / f'{page_name}.json'
    if not json_path.exists():
        print(f"✗ Evidence JSON not found: {json_path}")
        available = [p.stem for p in evidence_dir.glob('*.json')]
        print(f"  Available: {', '.join(available)}")
        sys.exit(1)

    data = json.loads(json_path.read_text(encoding='utf-8'))
    db = load_db()
    existing_dois = {s['doi'] for s in db['studies'] if s.get('doi')}
    added = 0

    for card in data.get('cards', []):
        for ref_html in card.get('studyRefs', []):
            refs = re.findall(r'\[(\d+)\]\s*(.*?)(?=<\/div>)', ref_html, re.DOTALL)
            for ref_num, ref_text in refs:
                # Extract DOI
                doi_match = re.search(r'href="https://doi\.org/([^"]+)"', ref_text)
                doi = doi_match.group(1) if doi_match else None
                if doi and doi in existing_dois:
                    continue

                # Extract authors
                clean = re.sub(r'<[^>]+>', '', ref_text).strip()
                authors_match = re.search(r'^(.*?)\s*\(\d{4}\)', clean)
                authors = authors_match.group(1).strip().rstrip(',').strip() if authors_match else ''
                authors = authors.replace('&amp;', '&')

                year_match = re.search(r'\((\d{4})\)', clean)
                year = int(year_match.group(1)) if year_match else 2020

                title_match = re.search(r'"([^"]+)"', ref_text)
                title = title_match.group(1) if title_match else ''

                sid = make_study_id(authors, year)
                db['studies'].append({
                    'study_id': sid, 'authors': authors,
                    'authors_short': authors if len(authors) < 30 else authors[:30],
                    'pub_year': year, 'title': title, 'doi': doi,
                    'url': f'https://doi.org/{doi}' if doi else '',
                    'study_type': 'Study', 'quality_score': 0,
                    'is_landmark': False, 'direction': '+',
                    'relevance': 1.0, 'final_score': 0,
                    'key_finding': '', 'status': 'active',
                    'verified_date': datetime.now().strftime('%Y-%m'),
                    'added_date': datetime.now().strftime('%Y-%m-%d'),
                    'notes': f'Auto-imported from {page_name}.json card {card["id"]}',
                })
                if doi:
                    existing_dois.add(doi)
                added += 1
                print(f"  + {authors} ({year}) from {card['id']}")

    save_db(db)
    print(f"\n✓ Imported {added} new studies from {page_name}.json")


# ─────────────────────────────────────────────
# CLI entry point
# ─────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description='LongevityPath Evidence Registry',
        formatter_class=argparse.RawDescriptionHelpFormatter, epilog=__doc__
    )
    sub = parser.add_subparsers(dest='command', help='Available commands')

    sub.add_parser('init', help='Create empty studies.json')
    p_import = sub.add_parser('import-md', help='Import from study-registry.md')
    p_import.add_argument('file', help='Path to study-registry.md')
    p_query = sub.add_parser('query', help='Query studies by claim')
    p_query.add_argument('claim', help='Claim tag (e.g., protein→cancer)')
    p_query.add_argument('--dir', help='Filter by direction (+/−/±)')
    sub.add_parser('summary', help='Print Claim Summary Index')
    sub.add_parser('export-summary', help='Write Claim Summary Index to study-registry.md')
    p_refs = sub.add_parser('export-refs', help='Export builder-ready study refs')
    p_refs.add_argument('claim', help='Claim tag or "all"')
    p_refs.add_argument('--json', action='store_true', help='Output as JSON')
    sub.add_parser('stale', help='List studies needing re-verification')
    p_gaps = sub.add_parser('gaps', help='Show claims with weak or missing evidence')
    p_gaps.add_argument('--category', help='Filter by category (sleep, nutrition, etc.)')
    p_supersede = sub.add_parser('supersede', help='Mark a study as superseded')
    p_supersede.add_argument('old_id', help='Study ID to mark as superseded')
    p_supersede.add_argument('new_id', help='Study ID of the replacement')
    sub.add_parser('enrich-all', help='List studies with incomplete fields')
    p_xlsx = sub.add_parser('export-xlsx', help='Export studies to xlsx')
    p_xlsx.add_argument('category', nargs='?', help='Category to export (or all)')
    p_irefs = sub.add_parser('import-refs', help='Import studies from evidence page JSON')
    p_irefs.add_argument('page', help='Evidence page name (e.g., sleep, nutrition-principles)')
    sub.add_parser('add', help='Interactively add a new study')
    sub.add_parser('stats', help='Show statistics')
    p_vdoi = sub.add_parser('verify-dois', help='Batch-verify all DOIs against doi.org')
    p_vdoi.add_argument('--category', help='Filter by category (sleep, nutrition, etc.)')

    args = parser.parse_args()
    if not args.command:
        parser.print_help()
        sys.exit(1)

    commands = {
        'init': cmd_init, 'import-md': cmd_import_md, 'query': cmd_query,
        'summary': cmd_summary, 'export-summary': cmd_export_summary,
        'export-refs': cmd_export_refs, 'stale': cmd_stale, 'add': cmd_add,
        'stats': cmd_stats, 'gaps': cmd_gaps, 'supersede': cmd_supersede,
        'enrich-all': cmd_enrich_all, 'export-xlsx': cmd_export_xlsx,
        'import-refs': cmd_import_refs, 'verify-dois': cmd_verify_dois,
    }
    commands[args.command](args)


if __name__ == '__main__':
    main()
